#pip install openpyxl
import openpyxl
import sys
import os,os.path
import logging
#from pathlib import Path

#Входной каталог для перебора папок
input_dir = r'C:\Users\ubuntu\Documents\Кандидаты2\\'
#открываем книги для записи
#wb1 - выгрузка анкеты из Старк
wb1 = openpyxl.load_workbook(input_dir+'outputable1.xlsx', data_only=True)
#wb2 - новое заключение Эксел
wb2 = openpyxl.load_workbook(input_dir+'outputable2.xlsx', data_only=True)
#wb3 - результат проверки роботом Эксел
wb3 = openpyxl.load_workbook(input_dir+'outputable3.xlsx', data_only=True)
#wb4 - анкета  в Эксель для робота
wb4 = openpyxl.load_workbook(input_dir+'outputable4.xlsx', data_only=True)
#Перебираем рекурсивно каталоги
for subdir, dirs, files in os.walk(input_dir):
    for file in files:
        #ищем файлы с таблицами
        if file.endswith(".xlsx") or file.endswith(".xlsm"):
            in_file = os.path.join(subdir, file)
            #открываем книгу для чтения
            try:
                wb = openpyxl.load_workbook(in_file, read_only=True)
                #берем активный лист
                ws = wb.active
                #проверяем какой структуры таблица на разборке, если анкета из Старк, то разбираем
                if ws['B1'].value == "Организация":
                    #считываем данные из группы ячеек и заносим в список
                    cell_range1 = ws['A3':'AE3']
                    cell_lst1 = []
                    for cell in cell_range1:
                        for c in cell:
                            cell_lst1.append(c.value)
                    cell_lst1.append(subdir)
                    #берем активный лист выходной книги
                    ws1 = wb1.active
                    #записываем значения в ячейки
                    ws1.append(cell_lst1)
                    #закрываем книгу для чтения 
                    wb.close()
                #проверяем какой структуры таблица на разборке, если новое заключение Эксел, то разбираем
                elif ws['A1'].value == "Заключение":
                    #считываем данные из групп и ячеек и заносим в список
                    col_range2 = ws['C4':'C8']
                    col_lst2 = []
                    for cell in col_range2:
                        for c in cell:
                            col_lst2.append(c.value)
                    sp = ws['C9'].value
                    col_lst2.append(sp)
                    np = ws['D9'].value
                    col_lst2.append(np)
                    dvp = ws['E9'].value
                    col_lst2.append(dvp)
                    inn = ws['C10'].value
                    col_lst2.append(inn)
                    per_work1 = ws['C11'].value
                    col_lst2.append(per_work1)
                    work1 = ws['D11'].value
                    col_lst2.append(work1)
                    per_work2 = ws['C12'].value
                    col_lst2.append(per_work2)
                    work2 = ws['D12'].value
                    col_lst2.append(work2)
                    per_work3 = ws['C13'].value
                    col_lst2.append(per_work3)
                    work3 = ws['D13'].value
                    col_lst2.append(work3)
                    col_range2 = ws['C14':'C30']
                    for cell in col_range2:
                        for c in cell:
                            col_lst2.append(c.value)
                    col_lst2.append(subdir)
                    #берем активный лист выходной книги
                    ws2 = wb2.active
                    #записываем значения в ячейки
                    ws2.append(col_lst2)
                    #закрываем книгу для чтения 
                    wb.close()
                #проверяем какой структуры таблица на разборке, если результат работы Робота, то разбираем
                elif ws['A3'].value == "Вакансия":
                    #считываем данные из группы ячеек и заносим в список
                    col_range3 = ws['B3':'B30']
                    col_lst3 = []
                    for cell in col_range3:
                        for c in cell:
                            col_lst3.append(c.value)
                    col_lst3.append(subdir)
                    #берем активный лист выходной книги
                    ws3 = wb3.active
                    #записываем значения в ячейки
                    ws3.append(col_lst3)
                    #закрываем книгу для чтения 
                    wb.close()
                #проверяем какой структуры таблица на разборке, если старая анкета Эксел, то разбираем
                elif ws['A1'].value == "Анкета":
                    cell_lst4 = []
                    #считываем значение ячейки
                    fio = ws['B3'].value
                    cell_lst4.append(fio)
                    dr = ws['B7'].value
                    cell_lst4.append(dr)
                    mr = ws['E7'].value
                    cell_lst4.append(mr)
                    sp = ws['B9'].value
                    cell_lst4.append(sp)
                    np = ws['E9'].value
                    cell_lst4.append(np)
                    dv = ws['B10'].value
                    cell_lst4.append(dv)
                    kv = ws['B11'].value
                    cell_lst4.append(kv)
                    inn = ws['B4'].value
                    cell_lst4.append(inn)
                    snils = ws['E4'].value
                    cell_lst4.append(snils)
                    region = ws['E18'].value
                    cell_lst4.append(region)
                    city = ws['B19'].value
                    cell_lst4.append(city)
                    street = ws['E19'].value
                    cell_lst4.append(street)
                    build = ws['B20'].value
                    cell_lst4.append(build)
                    flat = ws['E20'].value
                    cell_lst4.append(flat)
                    phone = ws['B24'].value
                    cell_lst4.append(phone)
                    mail = ws['E25'].value
                    cell_lst4.append(mail)
                    #место работы
                    last_work = ws['B36'].value
                    cell_lst4.append(last_work)
                    start_work = ws['B37'].value
                    cell_lst4.append(start_work)
                    end_work = ws['E37'].value
                    cell_lst4.append(end_work)
                    staff = ws['B40'].value
                    cell_lst4.append(staff)
                    #место работы
                    last_work2 = ws['B44'].value
                    cell_lst4.append(last_work2)
                    start_work2 = ws['B45'].value
                    cell_lst4.append(start_work2)
                    end_work2 = ws['E45'].value
                    cell_lst4.append(end_work2)
                    staff2 = ws['B48'].value
                    cell_lst4.append(staff2)
                    #место работы
                    last_work1 = ws['B52'].value
                    cell_lst4.append(last_work1)
                    start_work1 = ws['B53'].value
                    cell_lst4.append(start_work1)
                    end_work1 = ws['E53'].value
                    cell_lst4.append(end_work1)
                    staff1 = ws['B56'].value
                    cell_lst4.append(staff1)
                    cell_lst4.append(subdir)
                    #берем активный лист выходной книги
                    ws4 = wb4.active
                    #записываем значения в ячейки
                    ws4.append(cell_lst4)
                    #закрываем книгу для чтения 
                    wb.close()
            except OSError:
                logging.exception('')
                #print ("")
#сохраняем книгу для записи
wb1.save(input_dir+'outputable1.xlsx')
wb2.save(input_dir+'outputable2.xlsx')
wb3.save(input_dir+'outputable3.xlsx')
wb4.save(input_dir+'outputable4.xlsx')
#wb1.close()